<<<<<<< HEAD
import xlrd
import xlwt
import xlsxwriter
import csv
import os
import random
from random import randrange
from random import randint
import pandas as pd
from openpyxl import workbook
from openpyxl import Workbook
from openpyxl import load_workbook
from csv import writer
from collections import OrderedDict, defaultdict
import statistics as st
from decimal import *
from random import shuffle
from itertools import groupby
import sys

_thisDir = os.path.dirname(os.path.abspath(__file__))
os.chdir(_thisDir)

### *********************************************************************************** ###
### 									*** CONTENT ***									###
### *********************************************************************************** ###

## list of function

## complicated and not so useful
# 				copyResultsPerf
# 				pilotA_ts
# 				copyResultsPerf_Training
# 				pilotA_tr
# 				copyResultsPerf_Training_5bl
# 				pilotA_tr_BL5

## more simple, but limited
# 				swipe_ts
# 				swipe_tr







## *******************														   ***************** ##
## *** 				* SCRIPT FOR AUTOMATIOZATION OF WRITING RESULTS PSYCHOPY *				 *** ##
## *******************													       ***************** ##



## variables to extract from 
## key_resp_test.corr
## pairType


## variables to write to
## token_{}.format(participantNumber)
## Bl_1 - Bl_5



## PATH HANDLING


## geting a list of files in particular directory

## maybe some useful functions
# print(os.path.dirname(os.path.realpath(Testing)))
# print(sys.argv[0])

## this directory path
thisDirPath = os.path.dirname(__file__)

## path for wake testing and sleep testing
path_W_Test = os.listdir(thisDirPath + '\\Pilot\\Psychopy\\Wake\\Testing')
path_S_Test = os.listdir(thisDirPath + '\\Pilot\\Psychopy\\Sleep\\Testing')


path_W_Train = os.listdir(thisDirPath + '\\Pilot\\Psychopy\\Wake\\Training')
path_S_Train = os.listdir(thisDirPath + '\\Pilot\\Psychopy\\Sleep\\Training')



## example of one file
token6 = '6_inferenceTesting_v1_2020-06-09_16h48.42.324.csv'    ## written on col1 = 1, col2 = 2






## --------------------- *** function for exporting testing data *** --------------------- ##

def copyResultsPerf(fileName, col1, col2, col3, col4, col5, cond, sheet=0, token=0):
	""" function to copy info on pairType and corr response from output file in psychopy to my result file """
	""" filename = file name without any other signs, col1 = for pair type where to write in result doc, """
	""" col2 = for corr response, cond can be S or W string either for S - sleep or W - wake, """
	""" sheet = which sheet to be written at, must be interger, it starts with 0, but columns start with 1 """

	## path to testing in wake and sleep
	pathW = 'Pilot/Psychopy/Wake/Testing'
	pathS = 'Pilot/Psychopy/Sleep/Testing'

	data_path_W = 'dataSetsFromPy/pilotA_wake_ts.xlsx'
	data_path_S = 'dataSetsFromPy/pilotA_sleep_ts.xlsx'

	## specifying the condition for which the data are read
	cond_Var = str()
	if 'W' == str(cond):
		cond_Var = pathW
		loadWB = data_path_W
		print(cond)
	elif 'S' == str(cond):
		cond_Var = pathS
		loadWB = data_path_S
		print(cond)

	# read csv results either for wake or sleep condition
	results_W_Test = pd.read_csv(cond_Var + '/' + str(fileName))
	results_W_Test_both = results_W_Test.loc[:, ['pairType', 'letterPos1', 'letterPos2', 'key_resp_test.corr', 'key_resp_test.rt']]

	## open my result file specified below <- if I use some other, then adjust the function to make it variable
	wb = load_workbook(loadWB)
	sheets = wb.sheetnames
	Sheet3 = wb[sheets[int(sheet)]]    ## specifying which sheet to use, that should change as well

	row = 0
	for c in range(len(results_W_Test_both)):
		# print(c)
		Sheet3.cell(row = row+2, column = int(col1)).value = results_W_Test_both.loc[c, 'pairType']
		Sheet3.cell(row = row+2, column = int(col2)).value = results_W_Test_both.loc[c, 'letterPos1']
		Sheet3.cell(row = row+2, column = int(col3)).value = results_W_Test_both.loc[c, 'letterPos2']
		Sheet3.cell(row = row+2, column = int(col4)).value = results_W_Test_both.loc[c, 'key_resp_test.corr']
		Sheet3.cell(row = row+2, column = int(col5)).value = results_W_Test_both.loc[c, 'key_resp_test.rt']
		row += 1

	Sheet3.cell(row = 1, column = col1).value = 'pairType_' + str(cond) + '_' + str(token)
	Sheet3.cell(row = 1, column = col2).value = 'letterPos1_' + str(cond) + '_' + str(token)
	Sheet3.cell(row = 1, column = col3).value = 'letterPos2_' + str(cond) + '_' + str(token)
	Sheet3.cell(row = 1, column = col4).value = 'key_resp_test.corr_' + str(cond) + '_' + str(token)
	Sheet3.cell(row = 1, column = col5).value = 'key_resp_test.rt_' + str(cond) + '_' + str(token)




	wb.save(loadWB)


# copyResultsPerf('6_inferenceTesting_v1_2020-06-09_16h48.42.324.csv', 1, 2, 3, 4, 5, 'W', token=6)
# copyResultsPerf('9_inferenceTesting_v1_2020-06-16_17h11.10.508.csv', 6, 7, 8, 9, 10, 'W', token=9)



### ---------------------------- *** FUNCTION FOR SWEAPING THE DATA SETS *** ----------------------------- ###

def pilotA_ts(cond):
# 	""" create csv file with testing data in wake condition """

	## path to testing in wake and sleep
	pathW = '\\Pilot\\Psychopy\\Wake\\Testing\\'
	pathS = '\\Pilot\\Psychopy\\Sleep\\Testing\\'

		## specifying the condition for which the data are read
	cond_Var = str()
	if 'W' == str(cond):
		cond_Var = pathW
		print(cond)
	elif 'S' == str(cond):
		cond_Var = pathS
		print(cond)


	## counters and lists holders
	token_temp = 0				## used in the sorting on one digit and two digits - two digits
	token_temp2 = 0				## one digit

	temp_files_oneDigit = []	## list holder for files with one digit
	temp_files_twoDigit = []	## list holder for files with two digit

	token_oneDigit = []			## list holder for tokens in one digit
	token_twoDigit = []			## list holder for tokens in two digit


	## initiate for loop for sorting on one digit and two digit
	for file in os.listdir(thisDirPath + str(cond_Var)):

		
		if file[1] == '_':							## equal to dash on position 2 - that's why 1 digit
			token_temp2 = str(file[0])				## swipe first position in file - which is token digit
			temp_files_oneDigit.append(file)		## append file name to holder
			token_oneDigit.append(int(token_temp2)) ## append token to holder

		elif file[1] != '_':						## not equal to dash on position 2 - that's why two digit, there is a number
			token_temp = str(file[0]) + str(file[1])## swipe first two positions in the file name, is token two digit
			temp_files_twoDigit.append(file)		## append to file name holder for two digit
			token_twoDigit.append(int(token_temp))	## append to token holder for one digit

	## initiate columns, with one digit starting at 1
	col1 = 1
	col2 = 2
	col3 = 3
	col4 = 4
	col5 = 5	

	countToken_oneDigit = 0      ## initialize token counter entered then to token list as identificator of position
	
	## for loop for one digit list of files
	for file_oneDigit in temp_files_oneDigit:
		print(file_oneDigit)
		
		## applying function written above
		copyResultsPerf(file_oneDigit, col1, col2, col3, col4, col5, cond, token = token_oneDigit[countToken_oneDigit])
		col1 += 5					## increase each column by 5 with next incrementation
		col2 += 5
		col3 += 5
		col4 += 5
		col5 += 5

		countToken_oneDigit += 1	## increase token counter for one digit

		## variables for last state of columns from the first for loop, start of second loop, so it start writing
		## at the position where last loop ended
		col1_temp = col1
		col2_temp = col2
		col3_temp = col3
		col4_temp = col4
		col5_temp = col5

	countToken_twoDigit = 0			## initialize token counter for two digit entered then to token list as identificator of position

	## for loop for two digit list of files
	for file_twoDigit in temp_files_twoDigit:
		## applying function written above
		copyResultsPerf(file_twoDigit, col1_temp, col2_temp, col3_temp, col4_temp, col5_temp, cond, token = token_twoDigit[countToken_twoDigit])
		col1_temp += 5     ## taking above variables with end state of previous loop
		col2_temp += 5
		col3_temp += 5
		col4_temp += 5
		col5_temp += 5

		countToken_twoDigit += 1	## token counter for two digit






## --------------------- *** function for exporting training data *** --------------------- ##

def copyResultsPerf_Training(fileName, col1, col2, col3, col4, col5, cond, sheet=0, token=0):
	""" function to copy info on pairType and corr response from output file in psychopy to my result file """
	""" filename = file name without any other signs, col1 = for pair type where to write in result doc, """
	""" col2 = for corr response, cond can be S or W string either for S - sleep or W - wake, """
	""" sheet = which sheet to be written at, must be interger, it starts with 0, but columns start with 1 """

	""" now the set up has 5 blocks instead of 3, so change it afterwards """

	data_path_W = 'dataSetsFromPy/pilotA_wake_tr.xlsx'
	data_path_S = 'dataSetsFromPy/pilotA_sleep_tr.xlsx'

	## path to testing in wake and sleep
	pathW_Tr = 'Pilot/Psychopy/Wake/Training'
	pathS_Tr = 'Pilot/Psychopy/Sleep/Training'

	## specifying the condition for which the data are read
	cond_Var = str()
	if 'W' == str(cond):
		cond_Var = pathW_Tr
		loadWB = data_path_W
		print(cond)
	elif 'S' == str(cond):
		cond_Var = pathS_Tr
		loadWB = data_path_S
		print(cond)


	# read csv results either for wake or sleep condition
	results_W_Tr = pd.read_csv(cond_Var + '/' + str(fileName))
	results_W_Tr_both = results_W_Tr.loc[:, ['pairType', 'letterPos1', 'letterPos2', 'key_resp_Bl.corr', 'key_resp_Bl.rt', 'key_resp_Criterion.corr', 'key_resp_Criterion.rt']]

	## open my result file specified below <- if I use some other, then adjust the function to make it variable
	wb = load_workbook(loadWB)
	sheets = wb.sheetnames
	Sheet3 = wb[sheets[int(sheet)]]    ## specifying which sheet to use, that should change as well

	row = 0
	for c in range(len(results_W_Tr_both)):
		# print(c)

		Sheet3.cell(row = row+2, column = int(col1)).value = results_W_Tr_both.loc[c, 'pairType']
		Sheet3.cell(row = row+2, column = int(col2)).value = results_W_Tr_both.loc[c, 'letterPos1']
		Sheet3.cell(row = row+2, column = int(col3)).value = results_W_Tr_both.loc[c, 'letterPos2']
		Sheet3.cell(row = row+2, column = int(col4)).value = results_W_Tr_both.loc[c, 'key_resp_Bl.corr']
		Sheet3.cell(row = row+2, column = int(col5)).value = results_W_Tr_both.loc[c, 'key_resp_Bl.rt']

		row += 1


	for h in range(32, len(results_W_Tr_both)):   ## here this number need to change
		# print(h)

		Sheet3.cell(row = h+2, column = int(col4)).value = results_W_Tr_both.loc[h, 'key_resp_Criterion.corr']     ## changed to key correct, not number corr
		Sheet3.cell(row = h+2, column = int(col5)).value = results_W_Tr_both.loc[h, 'key_resp_Criterion.rt']



	Sheet3.cell(row = 1, column = col1).value = 'pairType_' + str(cond) + '_' + str(token)
	Sheet3.cell(row = 1, column = col2).value = 'letterPos1_' + str(cond) + '_' + str(token)
	Sheet3.cell(row = 1, column = col3).value = 'letterPos2_' + str(cond) + '_' + str(token)
	Sheet3.cell(row = 1, column = col4).value = 'key_resp_Bl.corr_' + str(cond) + '_' + str(token)
	Sheet3.cell(row = 1, column = col5).value = 'key_resp_Bl.rt_' + str(cond) + '_' + str(token)





	wb.save(loadWB)





### ---------------------------- *** FUNCTION FOR SWEAPING THE DATA SETS *** ----------------------------- ###

def pilotA_tr(cond):
# 	""" create csv file with testing data in wake condition """

	## path to testing in wake and sleep
	pathW = '\\Pilot\\Psychopy\\Wake\\Training\\'
	pathS = '\\Pilot\\Psychopy\\Sleep\\Training\\'

		## specifying the condition for which the data are read
	cond_Var = str()
	if 'W' == str(cond):
		cond_Var = pathW
		print(cond)
	elif 'S' == str(cond):
		cond_Var = pathS
		print(cond)


	## counters and lists holders
	token_temp = 0				## used in the sorting on one digit and two digits - two digits
	token_temp2 = 0				## one digit

	temp_files_oneDigit = []	## list holder for files with one digit
	temp_files_twoDigit = []	## list holder for files with two digit

	token_oneDigit = []			## list holder for tokens in one digit
	token_twoDigit = []			## list holder for tokens in two digit


	## initiate for loop for sorting on one digit and two digit
	for file in os.listdir(thisDirPath + str(cond_Var)):

		
		if file[1] == '_':							## equal to dash on position 2 - that's why 1 digit
			token_temp2 = str(file[0])				## swipe first position in file - which is token digit
			temp_files_oneDigit.append(file)		## append file name to holder
			token_oneDigit.append(int(token_temp2)) ## append token to holder

		elif file[1] != '_':						## not equal to dash on position 2 - that's why two digit, there is a number
			token_temp = str(file[0]) + str(file[1])## swipe first two positions in the file name, is token two digit
			temp_files_twoDigit.append(file)		## append to file name holder for two digit
			token_twoDigit.append(int(token_temp))	## append to token holder for one digit

	## initiate columns, with one digit starting at 1
	col1 = 1
	col2 = 2
	col3 = 3
	col4 = 4
	col5 = 5	

	countToken_oneDigit = 0      ## initialize token counter entered then to token list as identificator of position
	
	## for loop for one digit list of files
	for file_oneDigit in temp_files_oneDigit:
		
		## applying function written above
		copyResultsPerf_Training(file_oneDigit, col1, col2, col3, col4, col5, cond, token = token_oneDigit[countToken_oneDigit])
		col1 += 5					## increase each column by 5 with next incrementation
		col2 += 5
		col3 += 5
		col4 += 5
		col5 += 5

		countToken_oneDigit += 1	## increase token counter for one digit

		## variables for last state of columns from the first for loop, start of second loop, so it start writing
		## at the position where last loop ended
		col1_temp = col1
		col2_temp = col2
		col3_temp = col3
		col4_temp = col4
		col5_temp = col5

	countToken_twoDigit = 0			## initialize token counter for two digit entered then to token list as identificator of position

	## for loop for two digit list of files
	for file_twoDigit in temp_files_twoDigit:
		## applying function written above
		copyResultsPerf_Training(file_twoDigit, col1_temp, col2_temp, col3_temp, col4_temp, col5_temp, cond, token = token_twoDigit[countToken_twoDigit])
		col1_temp += 5     ## taking above variables with end state of previous loop
		col2_temp += 5
		col3_temp += 5
		col4_temp += 5
		col5_temp += 5

		countToken_twoDigit += 1	## token counter for two digit




### ---------------------------- *** FUNCTION FOR SWEAPING THE DATA SETS *** ----------------------------- ###
## FOR TRAINING WITH 5 BLOCKS

def copyResultsPerf_Training_5bl(fileName, col1, col2, col3, col4, col5, cond, sheet=0, token=0):
	""" adaptation of copyResultsPerf_Training function, but with 5 testing blocks """
	""" second loop must be corrected from where it starts """

	data_path_W = 'dataSetsFromPy/pilotA_wake_tr_BL5.xlsx'
	data_path_S = 'dataSetsFromPy/pilotA_sleep_tr_BL5.xlsx'

	## path to testing in wake and sleep
	pathW_Tr = 'Pilot/Psychopy/Wake/Training_BL5'
	pathS_Tr = 'Pilot/Psychopy/Sleep/Training_BL5'

	## specifying the condition for which the data are read
	cond_Var = str()
	if 'W' == str(cond):
		cond_Var = pathW_Tr
		loadWB = data_path_W
		print(cond)
	elif 'S' == str(cond):
		cond_Var = pathS_Tr
		loadWB = data_path_S
		print(cond)

	# read csv results either for wake or sleep condition
	results_W_Tr = pd.read_csv(cond_Var + '/' + str(fileName))
	results_W_Tr_both = results_W_Tr.loc[:, ['pairType', 'letterPos1', 'letterPos2', 'key_resp_Bl.corr', 'key_resp_Bl.rt', 'key_resp_Criterion.corr', 'key_resp_Criterion.rt']]

	## open my result file specified below <- if I use some other, then adjust the function to make it variable
	wb = load_workbook(loadWB)
	sheets = wb.sheetnames
	#Sheet3 = wb[sheets[int(sheet)]]    ## specifying which sheet to use, that should change as well
	Sheet3 = wb[sheets[0]]
	print(Sheet3)

	row = 0
	for c in range(len(results_W_Tr_both)):
		print(c)

		Sheet3.cell(row = row+2, column = int(col1)).value = results_W_Tr_both.loc[c, 'pairType']
		Sheet3.cell(row = row+2, column = int(col2)).value = results_W_Tr_both.loc[c, 'letterPos1']
		Sheet3.cell(row = row+2, column = int(col3)).value = results_W_Tr_both.loc[c, 'letterPos2']
		Sheet3.cell(row = row+2, column = int(col4)).value = results_W_Tr_both.loc[c, 'key_resp_Bl.corr']
		Sheet3.cell(row = row+2, column = int(col5)).value = results_W_Tr_both.loc[c, 'key_resp_Bl.rt']

		row += 1

	for h in range(54, len(results_W_Tr_both)):   ## here this number need to change - where 5 bl ends
		# print(h)

		Sheet3.cell(row = h+2, column = int(col4)).value = results_W_Tr_both.loc[h, 'key_resp_Criterion.corr']     ## changed to key correct, not number corr
		Sheet3.cell(row = h+2, column = int(col5)).value = results_W_Tr_both.loc[h, 'key_resp_Criterion.rt']


	Sheet3.cell(row = 1, column = col1).value = 'pairType_' + str(cond) + '_' + str(token)
	Sheet3.cell(row = 1, column = col2).value = 'letterPos1_' + str(cond) + '_' + str(token)
	Sheet3.cell(row = 1, column = col3).value = 'letterPos2_' + str(cond) + '_' + str(token)
	Sheet3.cell(row = 1, column = col4).value = 'key_resp_Bl.corr_' + str(cond) + '_' + str(token)
	Sheet3.cell(row = 1, column = col5).value = 'key_resp_Bl.rt_' + str(cond) + '_' + str(token)

	wb.save(loadWB)



# copyResultsPerf_Training_5bl('17_inferenceTraining_v1_2020-06-26_09h50.09.600.csv', 1, 2, 3, 4, 5, 'W', 17)





### ---------------------------- *** FUNCTION FOR SWEAPING THE DATA SETS *** ----------------------------- ###

def pilotA_tr_BL5(cond):
# 	""" create csv file with testing data in wake condition """
	""" this must be chaged as well, because it uses function for 5 blocks """

	## path to testing in wake and sleep
	pathW = '\\Pilot\\Psychopy\\Wake\\Training_BL5\\'
	pathS = '\\Pilot\\Psychopy\\Sleep\\Training_BL5\\'

		## specifying the condition for which the data are read
	cond_Var = str()
	if 'W' == str(cond):
		cond_Var = pathW
		print(cond)
	elif 'S' == str(cond):
		cond_Var = pathS
		print(cond)

	## counters and lists holders
	token_temp = 0				## used in the sorting on one digit and two digits - two digits
	token_temp2 = 0				## one digit

	temp_files_oneDigit = []	## list holder for files with one digit
	temp_files_twoDigit = []	## list holder for files with two digit

	token_oneDigit = []			## list holder for tokens in one digit
	token_twoDigit = []			## list holder for tokens in two digit

	## initiate for loop for sorting on one digit and two digit
	for file in os.listdir(thisDirPath + str(cond_Var)):

		
		if file[1] == '_':							## equal to dash on position 2 - that's why 1 digit
			token_temp2 = str(file[0])				## swipe first position in file - which is token digit
			temp_files_oneDigit.append(file)		## append file name to holder
			token_oneDigit.append(int(token_temp2)) ## append token to holder

		elif file[1] != '_':						## not equal to dash on position 2 - that's why two digit, there is a number
			token_temp = str(file[0]) + str(file[1])## swipe first two positions in the file name, is token two digit
			temp_files_twoDigit.append(file)		## append to file name holder for two digit
			token_twoDigit.append(int(token_temp))	## append to token holder for one digit

	## initiate columns, with one digit starting at 1
	col1 = 1
	col2 = 2
	col3 = 3
	col4 = 4
	col5 = 5	

	countToken_oneDigit = 0      ## initialize token counter entered then to token list as identificator of position
	
	## for loop for one digit list of files
	for file_oneDigit in temp_files_oneDigit:
		
		## applying function written above
		copyResultsPerf_Training_5bl(file_oneDigit, col1, col2, col3, col4, col5, cond, token = token_oneDigit[countToken_oneDigit])
		col1 += 5					## increase each column by 5 with next incrementation
		col2 += 5
		col3 += 5
		col4 += 5
		col5 += 5

		countToken_oneDigit += 1	## increase token counter for one digit

	## variables for last state of columns from the first for loop, start of second loop, so it start writing
	## at the position where last loop ended
	col1_temp = col1
	col2_temp = col2
	col3_temp = col3
	col4_temp = col4
	col5_temp = col5

	countToken_twoDigit = 0			## initialize token counter for two digit entered then to token list as identificator of position

	## for loop for two digit list of files
	for file_twoDigit in temp_files_twoDigit:
		## applying function written above
		copyResultsPerf_Training_5bl(file_twoDigit, col1_temp, col2_temp, col3_temp, col4_temp, col5_temp, cond, token = token_twoDigit[countToken_twoDigit])
		col1_temp += 5     ## taking above variables with end state of previous loop
		col2_temp += 5
		col3_temp += 5
		col4_temp += 5
		col5_temp += 5

		countToken_twoDigit += 1	## token counter for two digit



# print(thisDirPath + 'R/dataSets_R')

targetFile = '\\R\\dataSets_Results_ts\\'

#print(thisDirPath + targetFile + '7_inferenceTesting_v1_2020-06-10_16h02.45.099.csv')


#print(thisDirPath + '\\' + targetFile '\\' + '7_inferenceTesting_v1_2020-06-10_16h02.45.099.csv')


## 		------------- *** Swiping functions ***	------------			##

## more useful, simple function for swiping the directory
## the main directory is dataSets_Results_ts or _tr, depending whether it's training or testing
## always copy files to the directory from psychopy and make some backup somewhere
## for now only pilot data there - so clean it when testing starts
## outputs one csv sheet to dataSetsFromPy -- but that can be changed


def swipe_ts():
	""" Swipe files for testing from defined directory - dataSets_Results_ts """
	""" Copy files to dataSets_Results_ts from psychopy """

	totalDf = []														 ## init empty list

	for file in os.listdir(thisDirPath + '\\R\\dataSets_Results_ts'):    ## loop throught dataSets_Results_ts dir
		resultsFile = pd.read_csv(thisDirPath + targetFile + str(file))  ## read files
		totalDf.append(resultsFile)										 ## append the dfs to the empty list

	totalDf_2 = pd.concat(totalDf, sort=False)							 ## concatanate the dfs in one df
	pd.DataFrame(totalDf_2).to_csv("dataSetsFromPy\\ts_tot.csv")		 ## output csv to dataSetsFromPy - maybe adjust that



# targetFile_tr = '\\R\\dataSets_Results_tr\\'




def swipe_tr():
	""" Swipe files for training from defined directory - dataSets_Results_tr """
	""" Copy files to dataSets_Results_ts from psychopy """

	totalDf = []															## init empty list			

	for file in os.listdir(thisDirPath + '\\R\\dataSets_Results_tr'):		## loop throught dataSets_Results_ts dir
		resultsFile = pd.read_csv(thisDirPath + targetFile_tr + str(file))  ## read files
		totalDf.append(resultsFile)											## append the dfs to the empty list

	totalDf_2 = pd.concat(totalDf, sort=False)								## concatanate the dfs in one df
	pd.DataFrame(totalDf_2).to_csv("dataSetsFromPy\\tr_tot.csv")			## output csv to dataSetsFromPy - maybe adjust that



## applying the function, no need to add arguments
swipe_ts()
swipe_tr()









### --------------- RUNNING THE SCRIPTS



# ## creating empty excels for for data
# filesFromPy = ['pilotA_wake_ts.xlsx', 'pilotA_sleep_ts.xlsx', 'pilotA_wake_tr.xlsx', 'pilotA_sleep_tr.xlsx']
# for filFromPy in filesFromPy:
# 	wb = Workbook()
# 	wb.save('dataSetsFromPy/' + filFromPy)

# filesFromPy = ['pilotA_wake_tr_BL5.xlsx', 'pilotA_sleep_tr_BL5.xlsx']
# for filFromPy in filesFromPy:
# 	wb = Workbook()
# 	wb.save('dataSetsFromPy/' + filFromPy)



# ## applying the second function for pilot wake condition, TRAINING
# pilotA_tr('W')
# pilotA_tr('S')

# pilotA_tr_BL5('W')
# pilotA_tr_BL5('S')


# ## applying the second function for pilot wake condition, TESTING
# pilotA_ts('W')
# pilotA_ts('S')


















## --------- *** 		TESTING          *** -------------------- ##
## --------- *** entering data for sleep *** -------------------- ##

## already exported: condition from token 7 and 8
## sheet 6
# copyResultsPerf('7_inferenceTesting_v1_2020-06-10_16h02.45.099.csv', 1, 2, 3, 'S', 6, token=7)
# copyResultsPerf('8_inferenceTesting_v1_2020-06-12_17h01.39.320.csv', 4, 5, 6, 'S', 6, token=8)


## to become exported



## -------- *** entering data for wake condition *** ------------- ##

## already exported: from token 6, 9
## sheet 4
# copyResultsPerf('6_inferenceTesting_v1_2020-06-09_16h48.42.324.csv', 1, 2, 3, 4, 5, 'W', 0, token=6)
# copyResultsPerf('9_inferenceTesting_v1_2020-06-16_17h11.10.508.csv', 4, 5, 6, 'W', 4, token=9)
# copyResultsPerf('15_inferenceTesting_v1_2020-06-18_23h01.33.535.csv', 7, 8, 9, 'W', 4, token=15)
# copyResultsPerf('11_inferenceTesting_v1_2020-06-19_16h28.32.423.csv', 10, 11, 12, 'W', 4, token=11)

## to become exported




## --------- *** 		TRAINING          *** -------------------- ##

## --------- *** entering data for sleep *** -------------------- ##

## already exported: condition from token 7 and 8
## sheet 9

# copyResultsPerf_Training('7_inferenceTraining_v1_2020-06-09_16h12.08.358.csv', 1, 2, 3, 'S', 9, token=7)
# copyResultsPerf_Training('8_inferenceTraining_v1_2020-06-11_17h10.03.506.csv', 4, 5, 6, 'S', 9, token=8)



## -------- *** entering data for wake condition *** ------------- ##

## already exported: from token 6, 9
## sheet 10

# start at 19 with column
# copyResultsPerf_Training('9_inferenceTraining_v1_2020-06-16_17h03.40.181.csv', 10, 11, 12, 'W', 10, token=9)
# copyResultsPerf_Training('11_inferenceTraining_v1_2020-06-19_16h10.15.577.csv', 13, 14, 15, 'W', 10, token=11)
# copyResultsPerf_Training('15_inferenceTraining_v1_2020-06-18_22h53.02.111.csv', 16, 17, 18, 'W', 10, token=15)




























## -------------------  *** ARCHIVE *** ------------------------- ##






# ## path for wake testing and sleep testing
# path_W_Test = os.listdir(thisDirPath + '\\Pilot\\Psychopy\\Wake\\Testing')
# path_S_Test = os.listdir(thisDirPath + '\\Pilot\\Psychopy\\Sleep\\Testing')


# path_W_Train = os.listdir(thisDirPath + '\\Pilot\\Psychopy\\Wake\\Training')
# path_S_Train = os.listdir(thisDirPath + '\\Pilot\\Psychopy\\Sleep\\Training')


# print(thisDirPath)

# t = pd.read_csv(thisDirPath + '\\Pilot\\Psychopy\\Wake\\Testing\\' + '15_inferenceTesting_v1_2020-06-18_23h01.33.535.csv')
# print(t)


# def csv_ts_W():
# 	""" create csv file with testing data in wake condition """

# 	# counter = 0
# 	listOfDf = []

# 	for file in os.listdir(thisDirPath + '\\Pilot\\Psychopy\\Wake\\Testing\\'):

# 		temp = pd.read_csv(thisDirPath + '\\Pilot\\Psychopy\\Wake\\Testing\\' + file)
# 		temp_reduced = temp.iloc[:, ['pairType', 'letterPos1', 'letterPos2', 'key_resp_test.corr', 'key_resp_test.rt']]
# 		listOfDf.append(temp_reduced)


# 		# counter += 1


# listOfPyFiles = []
# for item in os.listdir(thisDirPath + '\\dataSetsFromPy\\'):
# 	listOfPyFiles.append(item)

# print(listOfPyFiles)

# # fil = 'pilotA_wake_tr.xlsx'

# for fil in listOfPyFiles:
# 	# print(fil)
# 	# print(thisDirPath)
# 	wb = load_workbook(thisDirPath + '\\dataSetsFromPy\\' + fil)
# 	# sheets = wb.sheetnames
	# print(sheets)
	# wb.remove(wb[sheets[0]])
	# print(sheets[0])
	# wb.create_sheet(wb[sheets[0]], 0)
	# wb.save(thisDirPath + '\\dataSetsFromPy\\' + fil)
	 # workbook.remove_sheet(0)


# token = 0
# token_temp = 0
# counter_fake_1 = 0
# for file in os.listdir(thisDirPath + '\\Pilot\\Psychopy\\Wake\\Testing\\'):
# 	# print(file)
# 	# print(file[0])

	
# 	if file[1] != '_':
# 		token = str(file[0]) + str(file[1])
# 		token_temp = int(token)
# 		counter_fake_1 += 1
# 		# print(counter_fake_1)

# 		col1 = 1
# 		col2 = 2
# 		col3 = 3
# 		col4 = 4
# 		col5 = 5

# 		for i in range(1, (counter_fake_1*5)+1):
# 			# print(i)
# 			col1 = i
# 			col2 = i + 1
# 			col3 = i + 2
# 			col4 = i + 3
			# col5 = i + 4


		# print(col1)
		# print(col2)
		# print(col3)
		# print(col4)
		# print(col5)

		# print(token_temp)
		# col1_two = col1
		# col2_two = col2
		# col3_two = col3
		# col4_two = col4
		# col5_two = col5



## creating empty lists for writing later
# Bl_1W_corr = []
# Bl_2W_corr = []
# Bl_3W_corr = []
# Bl_4W_corr = []
# Bl_5W_corr = []

# Bl_1W_type = []
# Bl_2W_type = []
# Bl_3W_type = []
# Bl_4W_type = []
# Bl_5W_type = []



# ## lodading the file
# with open(str(pathW) + '/6_inferenceTesting_v1_2020-06-09_16h48.42.324.csv', newline='') as csvfile:
# 	reader = csv.reader(csvfile, skipinitialspace=True)
# 	holderCorr = []
# 	holderPair = []
# 	for row in reader:
# 		# print(row[7])
# 		if row[7] != 'key_resp_test.corr':
# 			holderCorr.append(row[7])
# 			# holderCorr3 = row[7]
# 		if row[19] != 'pairType':
# 			holderPair.append(row[19])	



# ## spliting list on empty string using groupby
# holderCorr2 = [list(sub) for ele, sub in groupby(holderCorr, key = bool) if ele]
# holderPair2 = [list(sub) for ele, sub in groupby(holderPair, key = bool) if ele]

# # for i in holderCorr2:
# # 	print(i)

# # print(type(holderPair2[0]))

# ## very stupid solution to extract integers and save them in above var, .. yeah I'm tired
# for i in holderCorr2[0]:
# 	Bl_1W_corr.append(int(i))
# for i in holderCorr2[1]:
# 	Bl_2W_corr.append(int(i))
# for i in holderCorr2[2]:
# 	Bl_3W_corr.append(int(i))
# for i in holderCorr2[3]:
# 	Bl_4W_corr.append(int(i))
# for i in holderCorr2[4]:
# 	Bl_5W_corr.append(int(i))


# holderPair2[0]
# Bl_1W_type = holderPair2[0]
# Bl_2W_type = holderPair2[1]
# Bl_3W_type = holderPair2[2]
# Bl_4W_type = holderPair2[3]
# Bl_5W_type = holderPair2[4]


# ## convert to dataFrame
# Bl_1_both = pd.DataFrame(list(zip(Bl_1W_type, Bl_1W_corr)), columns=['pairType', 'corrAns'])
# ## sort based on type
# Bl_1_both = Bl_1_both.sort_values(by='pairType', ascending=True)


# Bl_2_both = pd.DataFrame(list(zip(Bl_2W_type, Bl_2W_corr)), columns=['pairType', 'corrAns'])
# Bl_2_both = Bl_2_both.sort_values(by='pairType', ascending=True)



# Bl_1W_corrSort = Bl_1_both.loc[:, 'corrAns']
# Bl_1W_corrSort = list(Bl_1W_corrSort)
# # print(Bl_1W_corrSort)


# Bl_2W_corrSort = Bl_2_both.loc[:, 'corrAns']
# Bl_2W_corrSort = list(Bl_2W_corrSort)

# wb = load_workbook('Results_Pilot_Adults.xlsx')
# sheets = wb.sheetnames
# Sheet3 = wb[sheets[3]]

# row = 0
# row2 = 23
# for c in Bl_1W_corrSort:
# 	Sheet3.cell(row = row+3, column = 2).value = c
# 	row += 1

# for c1 in Bl_2W_corrSort:
# 	Sheet3.cell(row = row2, column = 2).value = c1
# 	row2 += 1



# wb.save('Results_Pilot_Adults.xlsx')










=======
import xlrd
import xlwt
import xlsxwriter
import csv
import os
import random
from random import randrange
from random import randint
import pandas as pd
from openpyxl import workbook
from openpyxl import Workbook
from openpyxl import load_workbook
from csv import writer
from collections import OrderedDict, defaultdict
import statistics as st
from decimal import *
from random import shuffle
from itertools import groupby
import sys

_thisDir = os.path.dirname(os.path.abspath(__file__))
os.chdir(_thisDir)

### *********************************************************************************** ###
### 									*** CONTENT ***									###
### *********************************************************************************** ###

## list of function

## complicated and not so useful
# 				copyResultsPerf
# 				pilotA_ts
# 				copyResultsPerf_Training
# 				pilotA_tr
# 				copyResultsPerf_Training_5bl
# 				pilotA_tr_BL5

## more simple, but limited
# 				swipe_ts
# 				swipe_tr







## *******************														   ***************** ##
## *** 				* SCRIPT FOR AUTOMATIOZATION OF WRITING RESULTS PSYCHOPY *				 *** ##
## *******************													       ***************** ##



## variables to extract from 
## key_resp_test.corr
## pairType


## variables to write to
## token_{}.format(participantNumber)
## Bl_1 - Bl_5



## PATH HANDLING


## geting a list of files in particular directory

## maybe some useful functions
# print(os.path.dirname(os.path.realpath(Testing)))
# print(sys.argv[0])

## this directory path
thisDirPath = os.path.dirname(__file__)

## path for wake testing and sleep testing
path_W_Test = os.listdir(thisDirPath + '\\Pilot\\Psychopy\\Wake\\Testing')
path_S_Test = os.listdir(thisDirPath + '\\Pilot\\Psychopy\\Sleep\\Testing')


path_W_Train = os.listdir(thisDirPath + '\\Pilot\\Psychopy\\Wake\\Training')
path_S_Train = os.listdir(thisDirPath + '\\Pilot\\Psychopy\\Sleep\\Training')



## example of one file
token6 = '6_inferenceTesting_v1_2020-06-09_16h48.42.324.csv'    ## written on col1 = 1, col2 = 2






## --------------------- *** function for exporting testing data *** --------------------- ##

def copyResultsPerf(fileName, col1, col2, col3, col4, col5, cond, sheet=0, token=0):
	""" function to copy info on pairType and corr response from output file in psychopy to my result file """
	""" filename = file name without any other signs, col1 = for pair type where to write in result doc, """
	""" col2 = for corr response, cond can be S or W string either for S - sleep or W - wake, """
	""" sheet = which sheet to be written at, must be interger, it starts with 0, but columns start with 1 """

	## path to testing in wake and sleep
	pathW = 'Pilot/Psychopy/Wake/Testing'
	pathS = 'Pilot/Psychopy/Sleep/Testing'

	data_path_W = 'dataSetsFromPy/pilotA_wake_ts.xlsx'
	data_path_S = 'dataSetsFromPy/pilotA_sleep_ts.xlsx'

	## specifying the condition for which the data are read
	cond_Var = str()
	if 'W' == str(cond):
		cond_Var = pathW
		loadWB = data_path_W
		print(cond)
	elif 'S' == str(cond):
		cond_Var = pathS
		loadWB = data_path_S
		print(cond)

	# read csv results either for wake or sleep condition
	results_W_Test = pd.read_csv(cond_Var + '/' + str(fileName))
	results_W_Test_both = results_W_Test.loc[:, ['pairType', 'letterPos1', 'letterPos2', 'key_resp_test.corr', 'key_resp_test.rt']]

	## open my result file specified below <- if I use some other, then adjust the function to make it variable
	wb = load_workbook(loadWB)
	sheets = wb.sheetnames
	Sheet3 = wb[sheets[int(sheet)]]    ## specifying which sheet to use, that should change as well

	row = 0
	for c in range(len(results_W_Test_both)):
		# print(c)
		Sheet3.cell(row = row+2, column = int(col1)).value = results_W_Test_both.loc[c, 'pairType']
		Sheet3.cell(row = row+2, column = int(col2)).value = results_W_Test_both.loc[c, 'letterPos1']
		Sheet3.cell(row = row+2, column = int(col3)).value = results_W_Test_both.loc[c, 'letterPos2']
		Sheet3.cell(row = row+2, column = int(col4)).value = results_W_Test_both.loc[c, 'key_resp_test.corr']
		Sheet3.cell(row = row+2, column = int(col5)).value = results_W_Test_both.loc[c, 'key_resp_test.rt']
		row += 1

	Sheet3.cell(row = 1, column = col1).value = 'pairType_' + str(cond) + '_' + str(token)
	Sheet3.cell(row = 1, column = col2).value = 'letterPos1_' + str(cond) + '_' + str(token)
	Sheet3.cell(row = 1, column = col3).value = 'letterPos2_' + str(cond) + '_' + str(token)
	Sheet3.cell(row = 1, column = col4).value = 'key_resp_test.corr_' + str(cond) + '_' + str(token)
	Sheet3.cell(row = 1, column = col5).value = 'key_resp_test.rt_' + str(cond) + '_' + str(token)




	wb.save(loadWB)


# copyResultsPerf('6_inferenceTesting_v1_2020-06-09_16h48.42.324.csv', 1, 2, 3, 4, 5, 'W', token=6)
# copyResultsPerf('9_inferenceTesting_v1_2020-06-16_17h11.10.508.csv', 6, 7, 8, 9, 10, 'W', token=9)



### ---------------------------- *** FUNCTION FOR SWEAPING THE DATA SETS *** ----------------------------- ###

def pilotA_ts(cond):
# 	""" create csv file with testing data in wake condition """

	## path to testing in wake and sleep
	pathW = '\\Pilot\\Psychopy\\Wake\\Testing\\'
	pathS = '\\Pilot\\Psychopy\\Sleep\\Testing\\'

		## specifying the condition for which the data are read
	cond_Var = str()
	if 'W' == str(cond):
		cond_Var = pathW
		print(cond)
	elif 'S' == str(cond):
		cond_Var = pathS
		print(cond)


	## counters and lists holders
	token_temp = 0				## used in the sorting on one digit and two digits - two digits
	token_temp2 = 0				## one digit

	temp_files_oneDigit = []	## list holder for files with one digit
	temp_files_twoDigit = []	## list holder for files with two digit

	token_oneDigit = []			## list holder for tokens in one digit
	token_twoDigit = []			## list holder for tokens in two digit


	## initiate for loop for sorting on one digit and two digit
	for file in os.listdir(thisDirPath + str(cond_Var)):

		
		if file[1] == '_':							## equal to dash on position 2 - that's why 1 digit
			token_temp2 = str(file[0])				## swipe first position in file - which is token digit
			temp_files_oneDigit.append(file)		## append file name to holder
			token_oneDigit.append(int(token_temp2)) ## append token to holder

		elif file[1] != '_':						## not equal to dash on position 2 - that's why two digit, there is a number
			token_temp = str(file[0]) + str(file[1])## swipe first two positions in the file name, is token two digit
			temp_files_twoDigit.append(file)		## append to file name holder for two digit
			token_twoDigit.append(int(token_temp))	## append to token holder for one digit

	## initiate columns, with one digit starting at 1
	col1 = 1
	col2 = 2
	col3 = 3
	col4 = 4
	col5 = 5	

	countToken_oneDigit = 0      ## initialize token counter entered then to token list as identificator of position
	
	## for loop for one digit list of files
	for file_oneDigit in temp_files_oneDigit:
		print(file_oneDigit)
		
		## applying function written above
		copyResultsPerf(file_oneDigit, col1, col2, col3, col4, col5, cond, token = token_oneDigit[countToken_oneDigit])
		col1 += 5					## increase each column by 5 with next incrementation
		col2 += 5
		col3 += 5
		col4 += 5
		col5 += 5

		countToken_oneDigit += 1	## increase token counter for one digit

		## variables for last state of columns from the first for loop, start of second loop, so it start writing
		## at the position where last loop ended
		col1_temp = col1
		col2_temp = col2
		col3_temp = col3
		col4_temp = col4
		col5_temp = col5

	countToken_twoDigit = 0			## initialize token counter for two digit entered then to token list as identificator of position

	## for loop for two digit list of files
	for file_twoDigit in temp_files_twoDigit:
		## applying function written above
		copyResultsPerf(file_twoDigit, col1_temp, col2_temp, col3_temp, col4_temp, col5_temp, cond, token = token_twoDigit[countToken_twoDigit])
		col1_temp += 5     ## taking above variables with end state of previous loop
		col2_temp += 5
		col3_temp += 5
		col4_temp += 5
		col5_temp += 5

		countToken_twoDigit += 1	## token counter for two digit






## --------------------- *** function for exporting training data *** --------------------- ##

def copyResultsPerf_Training(fileName, col1, col2, col3, col4, col5, cond, sheet=0, token=0):
	""" function to copy info on pairType and corr response from output file in psychopy to my result file """
	""" filename = file name without any other signs, col1 = for pair type where to write in result doc, """
	""" col2 = for corr response, cond can be S or W string either for S - sleep or W - wake, """
	""" sheet = which sheet to be written at, must be interger, it starts with 0, but columns start with 1 """

	""" now the set up has 5 blocks instead of 3, so change it afterwards """

	data_path_W = 'dataSetsFromPy/pilotA_wake_tr.xlsx'
	data_path_S = 'dataSetsFromPy/pilotA_sleep_tr.xlsx'

	## path to testing in wake and sleep
	pathW_Tr = 'Pilot/Psychopy/Wake/Training'
	pathS_Tr = 'Pilot/Psychopy/Sleep/Training'

	## specifying the condition for which the data are read
	cond_Var = str()
	if 'W' == str(cond):
		cond_Var = pathW_Tr
		loadWB = data_path_W
		print(cond)
	elif 'S' == str(cond):
		cond_Var = pathS_Tr
		loadWB = data_path_S
		print(cond)


	# read csv results either for wake or sleep condition
	results_W_Tr = pd.read_csv(cond_Var + '/' + str(fileName))
	results_W_Tr_both = results_W_Tr.loc[:, ['pairType', 'letterPos1', 'letterPos2', 'key_resp_Bl.corr', 'key_resp_Bl.rt', 'key_resp_Criterion.corr', 'key_resp_Criterion.rt']]

	## open my result file specified below <- if I use some other, then adjust the function to make it variable
	wb = load_workbook(loadWB)
	sheets = wb.sheetnames
	Sheet3 = wb[sheets[int(sheet)]]    ## specifying which sheet to use, that should change as well

	row = 0
	for c in range(len(results_W_Tr_both)):
		# print(c)

		Sheet3.cell(row = row+2, column = int(col1)).value = results_W_Tr_both.loc[c, 'pairType']
		Sheet3.cell(row = row+2, column = int(col2)).value = results_W_Tr_both.loc[c, 'letterPos1']
		Sheet3.cell(row = row+2, column = int(col3)).value = results_W_Tr_both.loc[c, 'letterPos2']
		Sheet3.cell(row = row+2, column = int(col4)).value = results_W_Tr_both.loc[c, 'key_resp_Bl.corr']
		Sheet3.cell(row = row+2, column = int(col5)).value = results_W_Tr_both.loc[c, 'key_resp_Bl.rt']

		row += 1


	for h in range(32, len(results_W_Tr_both)):   ## here this number need to change
		# print(h)

		Sheet3.cell(row = h+2, column = int(col4)).value = results_W_Tr_both.loc[h, 'key_resp_Criterion.corr']     ## changed to key correct, not number corr
		Sheet3.cell(row = h+2, column = int(col5)).value = results_W_Tr_both.loc[h, 'key_resp_Criterion.rt']



	Sheet3.cell(row = 1, column = col1).value = 'pairType_' + str(cond) + '_' + str(token)
	Sheet3.cell(row = 1, column = col2).value = 'letterPos1_' + str(cond) + '_' + str(token)
	Sheet3.cell(row = 1, column = col3).value = 'letterPos2_' + str(cond) + '_' + str(token)
	Sheet3.cell(row = 1, column = col4).value = 'key_resp_Bl.corr_' + str(cond) + '_' + str(token)
	Sheet3.cell(row = 1, column = col5).value = 'key_resp_Bl.rt_' + str(cond) + '_' + str(token)





	wb.save(loadWB)





### ---------------------------- *** FUNCTION FOR SWEAPING THE DATA SETS *** ----------------------------- ###

def pilotA_tr(cond):
# 	""" create csv file with testing data in wake condition """

	## path to testing in wake and sleep
	pathW = '\\Pilot\\Psychopy\\Wake\\Training\\'
	pathS = '\\Pilot\\Psychopy\\Sleep\\Training\\'

		## specifying the condition for which the data are read
	cond_Var = str()
	if 'W' == str(cond):
		cond_Var = pathW
		print(cond)
	elif 'S' == str(cond):
		cond_Var = pathS
		print(cond)


	## counters and lists holders
	token_temp = 0				## used in the sorting on one digit and two digits - two digits
	token_temp2 = 0				## one digit

	temp_files_oneDigit = []	## list holder for files with one digit
	temp_files_twoDigit = []	## list holder for files with two digit

	token_oneDigit = []			## list holder for tokens in one digit
	token_twoDigit = []			## list holder for tokens in two digit


	## initiate for loop for sorting on one digit and two digit
	for file in os.listdir(thisDirPath + str(cond_Var)):

		
		if file[1] == '_':							## equal to dash on position 2 - that's why 1 digit
			token_temp2 = str(file[0])				## swipe first position in file - which is token digit
			temp_files_oneDigit.append(file)		## append file name to holder
			token_oneDigit.append(int(token_temp2)) ## append token to holder

		elif file[1] != '_':						## not equal to dash on position 2 - that's why two digit, there is a number
			token_temp = str(file[0]) + str(file[1])## swipe first two positions in the file name, is token two digit
			temp_files_twoDigit.append(file)		## append to file name holder for two digit
			token_twoDigit.append(int(token_temp))	## append to token holder for one digit

	## initiate columns, with one digit starting at 1
	col1 = 1
	col2 = 2
	col3 = 3
	col4 = 4
	col5 = 5	

	countToken_oneDigit = 0      ## initialize token counter entered then to token list as identificator of position
	
	## for loop for one digit list of files
	for file_oneDigit in temp_files_oneDigit:
		
		## applying function written above
		copyResultsPerf_Training(file_oneDigit, col1, col2, col3, col4, col5, cond, token = token_oneDigit[countToken_oneDigit])
		col1 += 5					## increase each column by 5 with next incrementation
		col2 += 5
		col3 += 5
		col4 += 5
		col5 += 5

		countToken_oneDigit += 1	## increase token counter for one digit

		## variables for last state of columns from the first for loop, start of second loop, so it start writing
		## at the position where last loop ended
		col1_temp = col1
		col2_temp = col2
		col3_temp = col3
		col4_temp = col4
		col5_temp = col5

	countToken_twoDigit = 0			## initialize token counter for two digit entered then to token list as identificator of position

	## for loop for two digit list of files
	for file_twoDigit in temp_files_twoDigit:
		## applying function written above
		copyResultsPerf_Training(file_twoDigit, col1_temp, col2_temp, col3_temp, col4_temp, col5_temp, cond, token = token_twoDigit[countToken_twoDigit])
		col1_temp += 5     ## taking above variables with end state of previous loop
		col2_temp += 5
		col3_temp += 5
		col4_temp += 5
		col5_temp += 5

		countToken_twoDigit += 1	## token counter for two digit




### ---------------------------- *** FUNCTION FOR SWEAPING THE DATA SETS *** ----------------------------- ###
## FOR TRAINING WITH 5 BLOCKS

def copyResultsPerf_Training_5bl(fileName, col1, col2, col3, col4, col5, cond, sheet=0, token=0):
	""" adaptation of copyResultsPerf_Training function, but with 5 testing blocks """
	""" second loop must be corrected from where it starts """

	data_path_W = 'dataSetsFromPy/pilotA_wake_tr_BL5.xlsx'
	data_path_S = 'dataSetsFromPy/pilotA_sleep_tr_BL5.xlsx'

	## path to testing in wake and sleep
	pathW_Tr = 'Pilot/Psychopy/Wake/Training_BL5'
	pathS_Tr = 'Pilot/Psychopy/Sleep/Training_BL5'

	## specifying the condition for which the data are read
	cond_Var = str()
	if 'W' == str(cond):
		cond_Var = pathW_Tr
		loadWB = data_path_W
		print(cond)
	elif 'S' == str(cond):
		cond_Var = pathS_Tr
		loadWB = data_path_S
		print(cond)

	# read csv results either for wake or sleep condition
	results_W_Tr = pd.read_csv(cond_Var + '/' + str(fileName))
	results_W_Tr_both = results_W_Tr.loc[:, ['pairType', 'letterPos1', 'letterPos2', 'key_resp_Bl.corr', 'key_resp_Bl.rt', 'key_resp_Criterion.corr', 'key_resp_Criterion.rt']]

	## open my result file specified below <- if I use some other, then adjust the function to make it variable
	wb = load_workbook(loadWB)
	sheets = wb.sheetnames
	#Sheet3 = wb[sheets[int(sheet)]]    ## specifying which sheet to use, that should change as well
	Sheet3 = wb[sheets[0]]
	print(Sheet3)

	row = 0
	for c in range(len(results_W_Tr_both)):
		print(c)

		Sheet3.cell(row = row+2, column = int(col1)).value = results_W_Tr_both.loc[c, 'pairType']
		Sheet3.cell(row = row+2, column = int(col2)).value = results_W_Tr_both.loc[c, 'letterPos1']
		Sheet3.cell(row = row+2, column = int(col3)).value = results_W_Tr_both.loc[c, 'letterPos2']
		Sheet3.cell(row = row+2, column = int(col4)).value = results_W_Tr_both.loc[c, 'key_resp_Bl.corr']
		Sheet3.cell(row = row+2, column = int(col5)).value = results_W_Tr_both.loc[c, 'key_resp_Bl.rt']

		row += 1

	for h in range(54, len(results_W_Tr_both)):   ## here this number need to change - where 5 bl ends
		# print(h)

		Sheet3.cell(row = h+2, column = int(col4)).value = results_W_Tr_both.loc[h, 'key_resp_Criterion.corr']     ## changed to key correct, not number corr
		Sheet3.cell(row = h+2, column = int(col5)).value = results_W_Tr_both.loc[h, 'key_resp_Criterion.rt']


	Sheet3.cell(row = 1, column = col1).value = 'pairType_' + str(cond) + '_' + str(token)
	Sheet3.cell(row = 1, column = col2).value = 'letterPos1_' + str(cond) + '_' + str(token)
	Sheet3.cell(row = 1, column = col3).value = 'letterPos2_' + str(cond) + '_' + str(token)
	Sheet3.cell(row = 1, column = col4).value = 'key_resp_Bl.corr_' + str(cond) + '_' + str(token)
	Sheet3.cell(row = 1, column = col5).value = 'key_resp_Bl.rt_' + str(cond) + '_' + str(token)

	wb.save(loadWB)



# copyResultsPerf_Training_5bl('17_inferenceTraining_v1_2020-06-26_09h50.09.600.csv', 1, 2, 3, 4, 5, 'W', 17)





### ---------------------------- *** FUNCTION FOR SWEAPING THE DATA SETS *** ----------------------------- ###

def pilotA_tr_BL5(cond):
# 	""" create csv file with testing data in wake condition """
	""" this must be chaged as well, because it uses function for 5 blocks """

	## path to testing in wake and sleep
	pathW = '\\Pilot\\Psychopy\\Wake\\Training_BL5\\'
	pathS = '\\Pilot\\Psychopy\\Sleep\\Training_BL5\\'

		## specifying the condition for which the data are read
	cond_Var = str()
	if 'W' == str(cond):
		cond_Var = pathW
		print(cond)
	elif 'S' == str(cond):
		cond_Var = pathS
		print(cond)

	## counters and lists holders
	token_temp = 0				## used in the sorting on one digit and two digits - two digits
	token_temp2 = 0				## one digit

	temp_files_oneDigit = []	## list holder for files with one digit
	temp_files_twoDigit = []	## list holder for files with two digit

	token_oneDigit = []			## list holder for tokens in one digit
	token_twoDigit = []			## list holder for tokens in two digit

	## initiate for loop for sorting on one digit and two digit
	for file in os.listdir(thisDirPath + str(cond_Var)):

		
		if file[1] == '_':							## equal to dash on position 2 - that's why 1 digit
			token_temp2 = str(file[0])				## swipe first position in file - which is token digit
			temp_files_oneDigit.append(file)		## append file name to holder
			token_oneDigit.append(int(token_temp2)) ## append token to holder

		elif file[1] != '_':						## not equal to dash on position 2 - that's why two digit, there is a number
			token_temp = str(file[0]) + str(file[1])## swipe first two positions in the file name, is token two digit
			temp_files_twoDigit.append(file)		## append to file name holder for two digit
			token_twoDigit.append(int(token_temp))	## append to token holder for one digit

	## initiate columns, with one digit starting at 1
	col1 = 1
	col2 = 2
	col3 = 3
	col4 = 4
	col5 = 5	

	countToken_oneDigit = 0      ## initialize token counter entered then to token list as identificator of position
	
	## for loop for one digit list of files
	for file_oneDigit in temp_files_oneDigit:
		
		## applying function written above
		copyResultsPerf_Training_5bl(file_oneDigit, col1, col2, col3, col4, col5, cond, token = token_oneDigit[countToken_oneDigit])
		col1 += 5					## increase each column by 5 with next incrementation
		col2 += 5
		col3 += 5
		col4 += 5
		col5 += 5

		countToken_oneDigit += 1	## increase token counter for one digit

	## variables for last state of columns from the first for loop, start of second loop, so it start writing
	## at the position where last loop ended
	col1_temp = col1
	col2_temp = col2
	col3_temp = col3
	col4_temp = col4
	col5_temp = col5

	countToken_twoDigit = 0			## initialize token counter for two digit entered then to token list as identificator of position

	## for loop for two digit list of files
	for file_twoDigit in temp_files_twoDigit:
		## applying function written above
		copyResultsPerf_Training_5bl(file_twoDigit, col1_temp, col2_temp, col3_temp, col4_temp, col5_temp, cond, token = token_twoDigit[countToken_twoDigit])
		col1_temp += 5     ## taking above variables with end state of previous loop
		col2_temp += 5
		col3_temp += 5
		col4_temp += 5
		col5_temp += 5

		countToken_twoDigit += 1	## token counter for two digit



# print(thisDirPath + 'R/dataSets_R')

targetFile = '\\R\\dataSets_Results_ts\\'

#print(thisDirPath + targetFile + '7_inferenceTesting_v1_2020-06-10_16h02.45.099.csv')


#print(thisDirPath + '\\' + targetFile '\\' + '7_inferenceTesting_v1_2020-06-10_16h02.45.099.csv')


## 		------------- *** Swiping functions ***	------------			##

## more useful, simple function for swiping the directory
## the main directory is dataSets_Results_ts or _tr, depending whether it's training or testing
## always copy files to the directory from psychopy and make some backup somewhere
## for now only pilot data there - so clean it when testing starts
## outputs one csv sheet to dataSetsFromPy -- but that can be changed


def swipe_ts():
	""" Swipe files for testing from defined directory - dataSets_Results_ts """
	""" Copy files to dataSets_Results_ts from psychopy """

	totalDf = []														 ## init empty list

	for file in os.listdir(thisDirPath + '\\R\\dataSets_Results_ts'):    ## loop throught dataSets_Results_ts dir
		resultsFile = pd.read_csv(thisDirPath + targetFile + str(file))  ## read files
		totalDf.append(resultsFile)										 ## append the dfs to the empty list

	totalDf_2 = pd.concat(totalDf, sort=False)							 ## concatanate the dfs in one df
	pd.DataFrame(totalDf_2).to_csv("dataSetsFromPy\\ts_tot.csv")		 ## output csv to dataSetsFromPy - maybe adjust that



# targetFile_tr = '\\R\\dataSets_Results_tr\\'




def swipe_tr():
	""" Swipe files for training from defined directory - dataSets_Results_tr """
	""" Copy files to dataSets_Results_ts from psychopy """

	totalDf = []															## init empty list			

	for file in os.listdir(thisDirPath + '\\R\\dataSets_Results_tr'):		## loop throught dataSets_Results_ts dir
		resultsFile = pd.read_csv(thisDirPath + targetFile_tr + str(file))  ## read files
		totalDf.append(resultsFile)											## append the dfs to the empty list

	totalDf_2 = pd.concat(totalDf, sort=False)								## concatanate the dfs in one df
	pd.DataFrame(totalDf_2).to_csv("dataSetsFromPy\\tr_tot.csv")			## output csv to dataSetsFromPy - maybe adjust that



## applying the function, no need to add arguments
swipe_ts()
swipe_tr()









### --------------- RUNNING THE SCRIPTS



# ## creating empty excels for for data
# filesFromPy = ['pilotA_wake_ts.xlsx', 'pilotA_sleep_ts.xlsx', 'pilotA_wake_tr.xlsx', 'pilotA_sleep_tr.xlsx']
# for filFromPy in filesFromPy:
# 	wb = Workbook()
# 	wb.save('dataSetsFromPy/' + filFromPy)

# filesFromPy = ['pilotA_wake_tr_BL5.xlsx', 'pilotA_sleep_tr_BL5.xlsx']
# for filFromPy in filesFromPy:
# 	wb = Workbook()
# 	wb.save('dataSetsFromPy/' + filFromPy)



# ## applying the second function for pilot wake condition, TRAINING
# pilotA_tr('W')
# pilotA_tr('S')

# pilotA_tr_BL5('W')
# pilotA_tr_BL5('S')


# ## applying the second function for pilot wake condition, TESTING
# pilotA_ts('W')
# pilotA_ts('S')


















## --------- *** 		TESTING          *** -------------------- ##
## --------- *** entering data for sleep *** -------------------- ##

## already exported: condition from token 7 and 8
## sheet 6
# copyResultsPerf('7_inferenceTesting_v1_2020-06-10_16h02.45.099.csv', 1, 2, 3, 'S', 6, token=7)
# copyResultsPerf('8_inferenceTesting_v1_2020-06-12_17h01.39.320.csv', 4, 5, 6, 'S', 6, token=8)


## to become exported



## -------- *** entering data for wake condition *** ------------- ##

## already exported: from token 6, 9
## sheet 4
# copyResultsPerf('6_inferenceTesting_v1_2020-06-09_16h48.42.324.csv', 1, 2, 3, 4, 5, 'W', 0, token=6)
# copyResultsPerf('9_inferenceTesting_v1_2020-06-16_17h11.10.508.csv', 4, 5, 6, 'W', 4, token=9)
# copyResultsPerf('15_inferenceTesting_v1_2020-06-18_23h01.33.535.csv', 7, 8, 9, 'W', 4, token=15)
# copyResultsPerf('11_inferenceTesting_v1_2020-06-19_16h28.32.423.csv', 10, 11, 12, 'W', 4, token=11)

## to become exported




## --------- *** 		TRAINING          *** -------------------- ##

## --------- *** entering data for sleep *** -------------------- ##

## already exported: condition from token 7 and 8
## sheet 9

# copyResultsPerf_Training('7_inferenceTraining_v1_2020-06-09_16h12.08.358.csv', 1, 2, 3, 'S', 9, token=7)
# copyResultsPerf_Training('8_inferenceTraining_v1_2020-06-11_17h10.03.506.csv', 4, 5, 6, 'S', 9, token=8)



## -------- *** entering data for wake condition *** ------------- ##

## already exported: from token 6, 9
## sheet 10

# start at 19 with column
# copyResultsPerf_Training('9_inferenceTraining_v1_2020-06-16_17h03.40.181.csv', 10, 11, 12, 'W', 10, token=9)
# copyResultsPerf_Training('11_inferenceTraining_v1_2020-06-19_16h10.15.577.csv', 13, 14, 15, 'W', 10, token=11)
# copyResultsPerf_Training('15_inferenceTraining_v1_2020-06-18_22h53.02.111.csv', 16, 17, 18, 'W', 10, token=15)




























## -------------------  *** ARCHIVE *** ------------------------- ##






# ## path for wake testing and sleep testing
# path_W_Test = os.listdir(thisDirPath + '\\Pilot\\Psychopy\\Wake\\Testing')
# path_S_Test = os.listdir(thisDirPath + '\\Pilot\\Psychopy\\Sleep\\Testing')


# path_W_Train = os.listdir(thisDirPath + '\\Pilot\\Psychopy\\Wake\\Training')
# path_S_Train = os.listdir(thisDirPath + '\\Pilot\\Psychopy\\Sleep\\Training')


# print(thisDirPath)

# t = pd.read_csv(thisDirPath + '\\Pilot\\Psychopy\\Wake\\Testing\\' + '15_inferenceTesting_v1_2020-06-18_23h01.33.535.csv')
# print(t)


# def csv_ts_W():
# 	""" create csv file with testing data in wake condition """

# 	# counter = 0
# 	listOfDf = []

# 	for file in os.listdir(thisDirPath + '\\Pilot\\Psychopy\\Wake\\Testing\\'):

# 		temp = pd.read_csv(thisDirPath + '\\Pilot\\Psychopy\\Wake\\Testing\\' + file)
# 		temp_reduced = temp.iloc[:, ['pairType', 'letterPos1', 'letterPos2', 'key_resp_test.corr', 'key_resp_test.rt']]
# 		listOfDf.append(temp_reduced)


# 		# counter += 1


# listOfPyFiles = []
# for item in os.listdir(thisDirPath + '\\dataSetsFromPy\\'):
# 	listOfPyFiles.append(item)

# print(listOfPyFiles)

# # fil = 'pilotA_wake_tr.xlsx'

# for fil in listOfPyFiles:
# 	# print(fil)
# 	# print(thisDirPath)
# 	wb = load_workbook(thisDirPath + '\\dataSetsFromPy\\' + fil)
# 	# sheets = wb.sheetnames
	# print(sheets)
	# wb.remove(wb[sheets[0]])
	# print(sheets[0])
	# wb.create_sheet(wb[sheets[0]], 0)
	# wb.save(thisDirPath + '\\dataSetsFromPy\\' + fil)
	 # workbook.remove_sheet(0)


# token = 0
# token_temp = 0
# counter_fake_1 = 0
# for file in os.listdir(thisDirPath + '\\Pilot\\Psychopy\\Wake\\Testing\\'):
# 	# print(file)
# 	# print(file[0])

	
# 	if file[1] != '_':
# 		token = str(file[0]) + str(file[1])
# 		token_temp = int(token)
# 		counter_fake_1 += 1
# 		# print(counter_fake_1)

# 		col1 = 1
# 		col2 = 2
# 		col3 = 3
# 		col4 = 4
# 		col5 = 5

# 		for i in range(1, (counter_fake_1*5)+1):
# 			# print(i)
# 			col1 = i
# 			col2 = i + 1
# 			col3 = i + 2
# 			col4 = i + 3
			# col5 = i + 4


		# print(col1)
		# print(col2)
		# print(col3)
		# print(col4)
		# print(col5)

		# print(token_temp)
		# col1_two = col1
		# col2_two = col2
		# col3_two = col3
		# col4_two = col4
		# col5_two = col5



## creating empty lists for writing later
# Bl_1W_corr = []
# Bl_2W_corr = []
# Bl_3W_corr = []
# Bl_4W_corr = []
# Bl_5W_corr = []

# Bl_1W_type = []
# Bl_2W_type = []
# Bl_3W_type = []
# Bl_4W_type = []
# Bl_5W_type = []



# ## lodading the file
# with open(str(pathW) + '/6_inferenceTesting_v1_2020-06-09_16h48.42.324.csv', newline='') as csvfile:
# 	reader = csv.reader(csvfile, skipinitialspace=True)
# 	holderCorr = []
# 	holderPair = []
# 	for row in reader:
# 		# print(row[7])
# 		if row[7] != 'key_resp_test.corr':
# 			holderCorr.append(row[7])
# 			# holderCorr3 = row[7]
# 		if row[19] != 'pairType':
# 			holderPair.append(row[19])	



# ## spliting list on empty string using groupby
# holderCorr2 = [list(sub) for ele, sub in groupby(holderCorr, key = bool) if ele]
# holderPair2 = [list(sub) for ele, sub in groupby(holderPair, key = bool) if ele]

# # for i in holderCorr2:
# # 	print(i)

# # print(type(holderPair2[0]))

# ## very stupid solution to extract integers and save them in above var, .. yeah I'm tired
# for i in holderCorr2[0]:
# 	Bl_1W_corr.append(int(i))
# for i in holderCorr2[1]:
# 	Bl_2W_corr.append(int(i))
# for i in holderCorr2[2]:
# 	Bl_3W_corr.append(int(i))
# for i in holderCorr2[3]:
# 	Bl_4W_corr.append(int(i))
# for i in holderCorr2[4]:
# 	Bl_5W_corr.append(int(i))


# holderPair2[0]
# Bl_1W_type = holderPair2[0]
# Bl_2W_type = holderPair2[1]
# Bl_3W_type = holderPair2[2]
# Bl_4W_type = holderPair2[3]
# Bl_5W_type = holderPair2[4]


# ## convert to dataFrame
# Bl_1_both = pd.DataFrame(list(zip(Bl_1W_type, Bl_1W_corr)), columns=['pairType', 'corrAns'])
# ## sort based on type
# Bl_1_both = Bl_1_both.sort_values(by='pairType', ascending=True)


# Bl_2_both = pd.DataFrame(list(zip(Bl_2W_type, Bl_2W_corr)), columns=['pairType', 'corrAns'])
# Bl_2_both = Bl_2_both.sort_values(by='pairType', ascending=True)



# Bl_1W_corrSort = Bl_1_both.loc[:, 'corrAns']
# Bl_1W_corrSort = list(Bl_1W_corrSort)
# # print(Bl_1W_corrSort)


# Bl_2W_corrSort = Bl_2_both.loc[:, 'corrAns']
# Bl_2W_corrSort = list(Bl_2W_corrSort)

# wb = load_workbook('Results_Pilot_Adults.xlsx')
# sheets = wb.sheetnames
# Sheet3 = wb[sheets[3]]

# row = 0
# row2 = 23
# for c in Bl_1W_corrSort:
# 	Sheet3.cell(row = row+3, column = 2).value = c
# 	row += 1

# for c1 in Bl_2W_corrSort:
# 	Sheet3.cell(row = row2, column = 2).value = c1
# 	row2 += 1



# wb.save('Results_Pilot_Adults.xlsx')










>>>>>>> 023a6c01f955342df82c6a7bd62dabe436fa68c0
